"""Inspect GPKG files using sqlite3 (no geopandas needed) and read Helmet Excel."""
import sqlite3
import json
from pathlib import Path

ROOT = Path(__file__).parent.parent
ADB = ROOT / "data" / "adb"


def inspect_gpkg(path: Path):
    print(f"\n{'='*60}")
    print(f"GPKG: {path.name}  ({path.stat().st_size/1e6:.1f} MB)")
    print("="*60)
    con = sqlite3.connect(str(path))
    cur = con.cursor()

    # List layers (tables with geometry)
    try:
        cur.execute("SELECT table_name, column_name FROM gpkg_geometry_columns")
    except Exception:
        cur.execute("SELECT table_name, 'geom' FROM gpkg_contents WHERE data_type='features'")
    layers = cur.fetchall()
    print(f"Layers: {layers}")

    for table, geom_col in layers:
        cur.execute(f"SELECT COUNT(*) FROM \"{table}\"")
        n = cur.fetchone()[0]
        print(f"\nLayer '{table}': {n} rows")

        # Column names
        cur.execute(f"PRAGMA table_info(\"{table}\")")
        cols = [(row[1], row[2]) for row in cur.fetchall()]
        print(f"Columns ({len(cols)}):")
        for name, dtype in cols:
            if name != geom_col:
                print(f"  {name:<40} {dtype}")

        # Sample values from first row
        non_geom = [c for c, _ in cols if c != geom_col]
        quoted = ", ".join('"' + c + '"' for c in non_geom[:20])
        sample_sql = f'SELECT {quoted} FROM "{table}" LIMIT 1'
        try:
            cur.execute(sample_sql)
            row = cur.fetchone()
            if row:
                print("\nSample row:")
                for c, v in zip(non_geom[:20], row):
                    print(f"  {c:<40} = {repr(v)[:60]}")
        except Exception as e:
            print(f"  Sample error: {e}")

        # Speed limit distribution
        speed_cols = [c for c, _ in cols if "speed" in c.lower() or "Speed" in c]
        for sc in speed_cols[:3]:
            try:
                cur.execute(f'SELECT MIN("{sc}"), MAX("{sc}"), AVG("{sc}") FROM "{table}" WHERE "{sc}" IS NOT NULL AND "{sc}" > 0')
                r = cur.fetchone()
                if r and r[0] is not None:
                    print(f"\n  {sc}: min={r[0]:.1f}, max={r[1]:.1f}, avg={r[2]:.1f}")
            except Exception:
                pass

        # AnalysisStatus distribution
        status_cols = [c for c, _ in cols if "status" in c.lower() or "Status" in c or "Analysis" in c]
        for sc in status_cols[:2]:
            try:
                cur.execute(f'SELECT "{sc}", COUNT(*) FROM "{table}" GROUP BY "{sc}"')
                print(f"\n  {sc} distribution:")
                for val, cnt in cur.fetchall():
                    print(f"    {repr(val)}: {cnt}")
            except Exception:
                pass

    con.close()


def inspect_excel(path: Path):
    print(f"\n{'='*60}")
    print(f"EXCEL: {path.name}  ({path.stat().st_size/1e6:.2f} MB)")
    print("="*60)
    try:
        import pandas as pd
        sheets = pd.ExcelFile(str(path)).sheet_names
        print(f"Sheets: {sheets}")
        for sheet in sheets:
            df = pd.read_excel(str(path), sheet_name=sheet)
            print(f"\nSheet '{sheet}': {len(df)} rows x {len(df.columns)} cols")
            print(f"Columns: {list(df.columns)}")
            print(df.head(5).to_string())
    except ImportError:
        print("pandas not available — using openpyxl")
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(path))
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                print(f"\nSheet '{sheet}': {ws.max_row} rows x {ws.max_column} cols")
                for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                    print(f"  {row}")
        except ImportError:
            print("openpyxl also not available")


if __name__ == "__main__":
    for gpkg in sorted(ADB.glob("*.gpkg")):
        inspect_gpkg(gpkg)

    for xlsx in sorted(ADB.glob("*.xlsx")):
        inspect_excel(xlsx)
